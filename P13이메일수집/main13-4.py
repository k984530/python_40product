import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://news.v.daum.net/v/20220708183033551'

headers = {
    'User-Agent' : 'Mozilla/5.0',
    'Content-Type' : 'text/html; charset=utf-8'
}

response = requests.get(url, headers = headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))

print(results)

try:
    wb = load_workbook(r"P13이메일수집\email.xlsx", data_only = True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    
for result in results:
    sheet.append([result])
    
wb.save(r'P13이메일수집\email.xlsx')